"""
reports.py — Celery task for generating executive reports.

Architecture contract (CRITICAL):
  1. All numeric values are computed by app.services.aggregation — never by AI.
  2. The AI model receives only pre-computed scalar numbers and returns ONLY
     narrative text (summary + recommendations). It is explicitly prohibited
     from generating numeric values.
  3. The PDF/Excel/CSV are assembled from the `numbers` dict, not from AI output.
     AI output is injected only into designated narrative text sections.
"""
import io
import json
import logging
import os
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, List

from app.tasks.celery_app import celery_app
from app.database import SessionLocal
from app.config import settings
from app.models.report import Report
from app.services.aggregation import (
    compute_org_summary, compute_theme_breakdown, compute_monthly_trend,
    OrgSummary, ThemeBreakdown, MonthlyTrend,
)

logger = logging.getLogger("phishguard")

# ---------------------------------------------------------------------------
# AI Narrative helpers
# ---------------------------------------------------------------------------

NARRATIVE_SYSTEM_PROMPT = """\
You are a concise executive security report writer.
You will be given precise metrics computed from a security awareness training database.
Your job is to write ONLY:
  1. A 2-3 sentence executive summary interpreting what these statistics mean.
  2. Exactly 3 short, actionable recommendations (one sentence each).

STRICT RULES:
- Do NOT invent, estimate, or alter any numeric values. The numbers are provided by the system.
- Do NOT repeat the numeric values in your output — they are inserted separately.
- Return ONLY a valid JSON object: {"summary": string, "recommendations": [string, string, string]}
- No markdown, no preamble, no trailing text.
"""


def _build_numbers_dict(
    summary: OrgSummary,
    themes: List[ThemeBreakdown],
    trend: List[MonthlyTrend],
) -> dict:
    """Build the canonical dict of all numeric values that will appear in the report.
    This is the ONLY source of truth — not AI output."""
    return {
        "total_users": summary.total_users,
        "avg_risk_score": summary.avg_risk_score,
        "click_rate_pct": round(summary.click_rate * 100, 1),
        "report_rate_pct": round(summary.report_rate * 100, 1),
        "open_rate_pct": round(summary.open_rate * 100, 1),
        "total_emails_sent": summary.total_emails_sent,
        "total_clicks": summary.total_clicks,
        "total_reports": summary.total_reports,
        "highest_risk_dept_name": summary.highest_risk_dept.department_name if summary.highest_risk_dept else "N/A",
        "highest_risk_dept_score": summary.highest_risk_dept.avg_risk_score if summary.highest_risk_dept else 0.0,
        "lowest_risk_dept_name": summary.lowest_risk_dept.department_name if summary.lowest_risk_dept else "N/A",
        "lowest_risk_dept_score": summary.lowest_risk_dept.avg_risk_score if summary.lowest_risk_dept else 0.0,
        "top_theme": themes[0].theme if themes else "N/A",
        "top_theme_click_rate_pct": round(themes[0].click_rate * 100, 1) if themes else 0.0,
        "departments": [
            {
                "name": d.department_name,
                "avg_risk_score": d.avg_risk_score,
                "click_rate_pct": round(d.click_rate * 100, 1),
                "report_rate_pct": round(d.report_rate * 100, 1),
                "user_count": d.user_count,
            }
            for d in summary.departments
        ],
        "monthly_trend": [
            {
                "month": t.month,
                "avg_risk_score": t.avg_risk_score,
                "emails_sent": t.emails_sent,
                "emails_clicked": t.emails_clicked,
            }
            for t in trend
        ],
        "top_themes": [
            {
                "theme": t.theme,
                "click_rate_pct": round(t.click_rate * 100, 1),
                "total_sent": t.total_sent,
            }
            for t in themes[:5]
        ],
    }


def _get_ai_narrative(numbers: dict) -> dict:
    """Call AI for narrative text only. Falls back gracefully if AI is unavailable."""
    try:
        from app.services.ai_generator import AIGeneratorService, AIProviderError
        ai = AIGeneratorService()
        user_prompt = (
            "Here are the security awareness training statistics for this reporting period:\n"
            f"- Total employees: {numbers['total_users']}\n"
            f"- Average risk score: {numbers['avg_risk_score']}/100\n"
            f"- Click rate: {numbers['click_rate_pct']}%\n"
            f"- Report rate: {numbers['report_rate_pct']}%\n"
            f"- Total phishing emails sent: {numbers['total_emails_sent']}\n"
            f"- Most vulnerable department: {numbers['highest_risk_dept_name']} "
            f"(score {numbers['highest_risk_dept_score']})\n"
            f"- Best performing department: {numbers['lowest_risk_dept_name']}\n"
            f"- Highest-risk theme: {numbers['top_theme']} "
            f"({numbers['top_theme_click_rate_pct']}% click rate)\n\n"
            "Write the executive summary and exactly 3 recommendations as described."
        )
        raw = ai.provider.generate_completion(NARRATIVE_SYSTEM_PROMPT, user_prompt)
        # Strip markdown fences if present
        cleaned = raw.strip()
        if cleaned.startswith("```"):
            lines = cleaned.splitlines()
            cleaned = "\n".join(lines[1:-1] if lines[-1].startswith("```") else lines[1:]).strip()
        parsed = json.loads(cleaned)
        return {
            "summary": str(parsed.get("summary", "")),
            "recommendations": [str(r) for r in parsed.get("recommendations", [])],
        }
    except Exception as exc:
        logger.warning(f"AI narrative generation failed (non-fatal): {exc}")
        return {
            "summary": (
                "This executive summary could not be generated by the AI model. "
                "Please review the numeric statistics in the table below."
            ),
            "recommendations": [
                "Schedule targeted training for the highest-risk department.",
                "Review phishing simulation themes with the highest click rates.",
                "Increase reporting-rate targets and celebrate departments that report promptly.",
            ],
        }


# ---------------------------------------------------------------------------
# Chart rendering (matplotlib — server-side, no display)
# ---------------------------------------------------------------------------

def _render_dept_bar_chart(numbers: dict) -> bytes:
    """Render a horizontal bar chart of department risk scores. Returns PNG bytes."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.colors as mcolors

    depts = numbers.get("departments", [])
    if not depts:
        return b""

    names = [d["name"] for d in depts]
    scores = [d["avg_risk_score"] for d in depts]

    def score_color(s):
        if s >= 90: return "#10b981"
        if s >= 70: return "#eab308"
        if s >= 50: return "#f97316"
        return "#ef4444"

    colors = [score_color(s) for s in scores]

    fig, ax = plt.subplots(figsize=(7, max(2, len(names) * 0.5 + 1)))
    fig.patch.set_facecolor("#0f172a")
    ax.set_facecolor("#1e293b")
    ax.barh(names, scores, color=colors, height=0.6)
    ax.set_xlim(0, 100)
    ax.set_xlabel("Risk Score", color="#94a3b8", fontsize=9)
    ax.tick_params(colors="#94a3b8", labelsize=8)
    ax.spines[:].set_color("#334155")
    ax.set_title("Department Risk Scores", color="white", fontsize=11, pad=8)
    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=120, facecolor=fig.get_facecolor())
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _render_trend_line_chart(numbers: dict) -> bytes:
    """Render a monthly risk trend line chart. Returns PNG bytes."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    trend = [t for t in numbers.get("monthly_trend", []) if t.get("avg_risk_score") is not None]
    if not trend:
        return b""

    months = [t["month"] for t in trend]
    scores = [t["avg_risk_score"] for t in trend]

    fig, ax = plt.subplots(figsize=(7, 3))
    fig.patch.set_facecolor("#0f172a")
    ax.set_facecolor("#1e293b")
    ax.plot(months, scores, color="#6366f1", linewidth=2, marker="o", markersize=5, markerfacecolor="#818cf8")
    ax.set_ylim(0, 100)
    ax.set_xlabel("Month", color="#94a3b8", fontsize=9)
    ax.set_ylabel("Avg Risk Score", color="#94a3b8", fontsize=9)
    ax.tick_params(colors="#94a3b8", labelsize=8)
    ax.spines[:].set_color("#334155")
    ax.set_title("Monthly Risk Score Trend", color="white", fontsize=11, pad=8)
    plt.xticks(rotation=30, ha="right")
    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=120, facecolor=fig.get_facecolor())
    plt.close(fig)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# PDF generation (reportlab)
# ---------------------------------------------------------------------------

def _render_pdf(numbers: dict, narrative: dict, report_dir: Path) -> Path:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, HRFlowable
    )
    from reportlab.lib.enums import TA_LEFT, TA_CENTER

    pdf_path = report_dir / "executive_report.pdf"
    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    DARK = colors.HexColor("#0f172a")
    MED = colors.HexColor("#1e293b")
    BORDER = colors.HexColor("#334155")
    PRIMARY = colors.HexColor("#6366f1")
    WHITE = colors.white
    MUTED = colors.HexColor("#94a3b8")
    RED = colors.HexColor("#ef4444")
    GREEN = colors.HexColor("#10b981")

    h1 = ParagraphStyle("h1", fontSize=20, textColor=WHITE, spaceAfter=4, fontName="Helvetica-Bold", alignment=TA_CENTER)
    h2 = ParagraphStyle("h2", fontSize=13, textColor=PRIMARY, spaceAfter=6, fontName="Helvetica-Bold")
    body = ParagraphStyle("body", fontSize=9, textColor=MUTED, spaceAfter=4, leading=14)
    bullet = ParagraphStyle("bullet", fontSize=9, textColor=MUTED, spaceAfter=2, leftIndent=12, leading=14)

    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    story = []

    # --- Title ---
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph("PhishGuard — Executive Security Report", h1))
    story.append(Paragraph(f"Generated: {now_str}", ParagraphStyle("sub", fontSize=8, textColor=MUTED, alignment=TA_CENTER)))
    story.append(Spacer(1, 0.4 * cm))
    story.append(HRFlowable(width="100%", thickness=1, color=PRIMARY))
    story.append(Spacer(1, 0.4 * cm))

    # --- AI Narrative Summary ---
    story.append(Paragraph("Executive Summary", h2))
    story.append(Paragraph(narrative["summary"], body))
    story.append(Spacer(1, 0.3 * cm))

    # --- Key Metrics Table ---
    story.append(Paragraph("Key Metrics", h2))
    metric_data = [
        ["Metric", "Value"],
        ["Total Employees", str(numbers["total_users"])],
        ["Avg Organisation Risk Score", f"{numbers['avg_risk_score']} / 100"],
        ["Click Rate", f"{numbers['click_rate_pct']}%"],
        ["Report Rate", f"{numbers['report_rate_pct']}%"],
        ["Open Rate", f"{numbers['open_rate_pct']}%"],
        ["Total Simulations Sent", str(numbers["total_emails_sent"])],
        ["Total Link Clicks", str(numbers["total_clicks"])],
        ["Total Reports Filed", str(numbers["total_reports"])],
        ["Highest-Risk Department", f"{numbers['highest_risk_dept_name']} (score {numbers['highest_risk_dept_score']})"],
        ["Best-Performing Department", numbers["lowest_risk_dept_name"]],
        ["Top Phishing Theme", f"{numbers['top_theme']} ({numbers['top_theme_click_rate_pct']}% click rate)"],
    ]
    metric_table = Table(metric_data, colWidths=[9 * cm, 8 * cm])
    metric_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 1), (-1, -1), MED),
        ("TEXTCOLOR", (0, 1), (-1, -1), MUTED),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [MED, colors.HexColor("#263144")]),
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [MED, colors.HexColor("#162032")]),
    ]))
    story.append(metric_table)
    story.append(Spacer(1, 0.5 * cm))

    # --- Department Breakdown ---
    if numbers.get("departments"):
        story.append(Paragraph("Department Risk Breakdown", h2))
        dept_data = [["Department", "Risk Score", "Click Rate", "Report Rate", "Users"]]
        for d in numbers["departments"]:
            dept_data.append([
                d["name"],
                str(d["avg_risk_score"]),
                f"{d['click_rate_pct']}%",
                f"{d['report_rate_pct']}%",
                str(d["user_count"]),
            ])
        dept_table = Table(dept_data, colWidths=[6 * cm, 3 * cm, 3 * cm, 3 * cm, 2 * cm])
        dept_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 1), (-1, -1), MED),
            ("TEXTCOLOR", (0, 1), (-1, -1), MUTED),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [MED, colors.HexColor("#162032")]),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(dept_table)
        story.append(Spacer(1, 0.5 * cm))

    # --- Charts ---
    dept_chart = _render_dept_bar_chart(numbers)
    trend_chart = _render_trend_line_chart(numbers)

    if dept_chart:
        story.append(Paragraph("Department Risk Chart", h2))
        story.append(Image(io.BytesIO(dept_chart), width=15 * cm, height=max(4 * cm, len(numbers["departments"]) * 0.6 * cm)))
        story.append(Spacer(1, 0.4 * cm))

    if trend_chart:
        story.append(Paragraph("Monthly Risk Trend", h2))
        story.append(Image(io.BytesIO(trend_chart), width=15 * cm, height=6 * cm))
        story.append(Spacer(1, 0.4 * cm))

    # --- Recommendations ---
    story.append(HRFlowable(width="100%", thickness=1, color=BORDER))
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph("Actionable Recommendations", h2))
    for i, rec in enumerate(narrative.get("recommendations", []), 1):
        story.append(Paragraph(f"{i}. {rec}", bullet))
    story.append(Spacer(1, 0.3 * cm))

    # --- Footer disclaimer ---
    disclaimer = ParagraphStyle("disc", fontSize=7, textColor=colors.HexColor("#475569"), alignment=TA_CENTER)
    story.append(HRFlowable(width="100%", thickness=0.5, color=BORDER))
    story.append(Paragraph(
        "All numeric values in this report are computed directly from the PhishGuard database. "
        "AI-generated narrative text is limited to summary and recommendations only.",
        disclaimer,
    ))

    doc.build(story)
    return pdf_path


# ---------------------------------------------------------------------------
# Excel generation (openpyxl)
# ---------------------------------------------------------------------------

def _render_excel(numbers: dict, report_dir: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    xl_path = report_dir / "executive_report.xlsx"
    wb = Workbook()

    HEADER_FILL = PatternFill("solid", fgColor="6366f1")
    ROW_FILL_A = PatternFill("solid", fgColor="1e293b")
    ROW_FILL_B = PatternFill("solid", fgColor="162032")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    CELL_FONT = Font(color="94a3b8", size=9)

    thin = Side(style="thin", color="334155")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header_row(ws, row):
        for cell in ws[row]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    def style_data_row(ws, row, alt=False):
        for cell in ws[row]:
            cell.font = CELL_FONT
            cell.fill = ROW_FILL_B if alt else ROW_FILL_A
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    def auto_col_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # --- Sheet 1: Org Summary ---
    ws1 = wb.active
    ws1.title = "Org Summary"
    ws1.sheet_view.showGridLines = False
    ws1.append(["Metric", "Value"])
    style_header_row(ws1, 1)

    summary_rows = [
        ("Total Employees", numbers["total_users"]),
        ("Avg Risk Score", numbers["avg_risk_score"]),
        ("Click Rate (%)", numbers["click_rate_pct"]),
        ("Report Rate (%)", numbers["report_rate_pct"]),
        ("Open Rate (%)", numbers["open_rate_pct"]),
        ("Total Emails Sent", numbers["total_emails_sent"]),
        ("Total Clicks", numbers["total_clicks"]),
        ("Total Reports", numbers["total_reports"]),
        ("Highest-Risk Dept", numbers["highest_risk_dept_name"]),
        ("Highest-Risk Dept Score", numbers["highest_risk_dept_score"]),
        ("Best-Performing Dept", numbers["lowest_risk_dept_name"]),
        ("Best-Performing Dept Score", numbers["lowest_risk_dept_score"]),
        ("Top Theme", numbers["top_theme"]),
        ("Top Theme Click Rate (%)", numbers["top_theme_click_rate_pct"]),
    ]
    for i, (k, v) in enumerate(summary_rows, 2):
        ws1.append([k, v])
        style_data_row(ws1, i, alt=(i % 2 == 1))
    auto_col_width(ws1)

    # --- Sheet 2: Departments ---
    ws2 = wb.create_sheet("Departments")
    ws2.sheet_view.showGridLines = False
    ws2.append(["Department", "Risk Score", "Click Rate (%)", "Report Rate (%)", "User Count"])
    style_header_row(ws2, 1)
    for i, d in enumerate(numbers.get("departments", []), 2):
        ws2.append([d["name"], d["avg_risk_score"], d["click_rate_pct"], d["report_rate_pct"], d["user_count"]])
        style_data_row(ws2, i, alt=(i % 2 == 1))
    auto_col_width(ws2)

    # --- Sheet 3: Monthly Trend ---
    ws3 = wb.create_sheet("Monthly Trend")
    ws3.sheet_view.showGridLines = False
    ws3.append(["Month", "Avg Risk Score", "Emails Sent", "Emails Clicked"])
    style_header_row(ws3, 1)
    for i, t in enumerate(numbers.get("monthly_trend", []), 2):
        ws3.append([t["month"], t.get("avg_risk_score"), t["emails_sent"], t["emails_clicked"]])
        style_data_row(ws3, i, alt=(i % 2 == 1))
    auto_col_width(ws3)

    # --- Sheet 4: Top Themes ---
    ws4 = wb.create_sheet("Campaign Themes")
    ws4.sheet_view.showGridLines = False
    ws4.append(["Theme", "Click Rate (%)", "Total Sent"])
    style_header_row(ws4, 1)
    for i, t in enumerate(numbers.get("top_themes", []), 2):
        ws4.append([t["theme"], t["click_rate_pct"], t["total_sent"]])
        style_data_row(ws4, i, alt=(i % 2 == 1))
    auto_col_width(ws4)

    wb.save(str(xl_path))
    return xl_path


# ---------------------------------------------------------------------------
# CSV export (zip of multiple CSVs)
# ---------------------------------------------------------------------------

def _render_csv_zip(numbers: dict, report_dir: Path) -> Path:
    import csv

    zip_path = report_dir / "executive_report_csv.zip"
    with zipfile.ZipFile(str(zip_path), "w", zipfile.ZIP_DEFLATED) as zf:

        # org_summary.csv
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["metric", "value"])
        for k, v in [
            ("total_users", numbers["total_users"]),
            ("avg_risk_score", numbers["avg_risk_score"]),
            ("click_rate_pct", numbers["click_rate_pct"]),
            ("report_rate_pct", numbers["report_rate_pct"]),
            ("open_rate_pct", numbers["open_rate_pct"]),
            ("total_emails_sent", numbers["total_emails_sent"]),
            ("total_clicks", numbers["total_clicks"]),
            ("total_reports", numbers["total_reports"]),
            ("highest_risk_dept_name", numbers["highest_risk_dept_name"]),
            ("highest_risk_dept_score", numbers["highest_risk_dept_score"]),
            ("lowest_risk_dept_name", numbers["lowest_risk_dept_name"]),
            ("lowest_risk_dept_score", numbers["lowest_risk_dept_score"]),
        ]:
            w.writerow([k, v])
        zf.writestr("org_summary.csv", buf.getvalue())

        # departments.csv
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["department", "avg_risk_score", "click_rate_pct", "report_rate_pct", "user_count"])
        for d in numbers.get("departments", []):
            w.writerow([d["name"], d["avg_risk_score"], d["click_rate_pct"], d["report_rate_pct"], d["user_count"]])
        zf.writestr("departments.csv", buf.getvalue())

        # monthly_trend.csv
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["month", "avg_risk_score", "emails_sent", "emails_clicked"])
        for t in numbers.get("monthly_trend", []):
            w.writerow([t["month"], t.get("avg_risk_score"), t["emails_sent"], t["emails_clicked"]])
        zf.writestr("monthly_trend.csv", buf.getvalue())

        # top_themes.csv
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["theme", "click_rate_pct", "total_sent"])
        for t in numbers.get("top_themes", []):
            w.writerow([t["theme"], t["click_rate_pct"], t["total_sent"]])
        zf.writestr("top_themes.csv", buf.getvalue())

    return zip_path


# ---------------------------------------------------------------------------
# Main Celery task
# ---------------------------------------------------------------------------

@celery_app.task(
    name="app.tasks.reports.generate_executive_report",
    bind=True,
    max_retries=0,
)
def generate_executive_report(
    self,
    report_id: int,
    org_id: int,
    date_from_iso: Optional[str],
    date_to_iso: Optional[str],
    department_id: Optional[int],
    formats: List[str],
):
    """
    Celery task: generate executive report files and update the Report row.

    Steps:
      1. Mark report as running
      2. Run DB aggregation queries (numbers ONLY come from here)
      3. Ask AI for narrative text (summary + recommendations) based on numbers
      4. Render requested file formats from numbers + narrative
      5. Save files, update report row to completed
    """
    db = SessionLocal()
    report_dir = None
    try:
        # 1. Mark running
        report = db.query(Report).filter(Report.id == report_id).first()
        if not report:
            logger.error(f"Report {report_id} not found in DB")
            return

        report.status = "running"
        db.commit()

        # 2. Parse dates
        dt_from = datetime.fromisoformat(date_from_iso).replace(tzinfo=timezone.utc) if date_from_iso else None
        dt_to = datetime.fromisoformat(date_to_iso).replace(tzinfo=timezone.utc) if date_to_iso else None

        # 3. DB aggregation — ALL NUMBERS COME FROM HERE
        logger.info(f"[report:{report_id}] Running aggregation queries...")
        summary = compute_org_summary(db, org_id, dt_from, dt_to, department_id)
        themes = compute_theme_breakdown(db, org_id, dt_from, dt_to)
        trend = compute_monthly_trend(db, org_id, num_months=6)

        # Build the single canonical numbers dict
        numbers = _build_numbers_dict(summary, themes, trend)
        logger.info(f"[report:{report_id}] Aggregation complete: {numbers['total_users']} users, "
                    f"avg_risk={numbers['avg_risk_score']}, click_rate={numbers['click_rate_pct']}%")

        # 4. AI narrative — only text, never numbers
        logger.info(f"[report:{report_id}] Requesting AI narrative...")
        narrative = _get_ai_narrative(numbers)

        # 5. Render files
        report_dir = Path(settings.REPORTS_DIR) / str(report_id)
        report_dir.mkdir(parents=True, exist_ok=True)

        file_paths = {}
        if "pdf" in formats:
            logger.info(f"[report:{report_id}] Rendering PDF...")
            pdf_path = _render_pdf(numbers, narrative, report_dir)
            file_paths["pdf"] = str(pdf_path)

        if "excel" in formats:
            logger.info(f"[report:{report_id}] Rendering Excel...")
            xl_path = _render_excel(numbers, report_dir)
            file_paths["excel"] = str(xl_path)

        if "csv" in formats:
            logger.info(f"[report:{report_id}] Rendering CSV zip...")
            csv_path = _render_csv_zip(numbers, report_dir)
            file_paths["csv"] = str(csv_path)

        # 6. Update report row
        report.status = "completed"
        report.file_paths = file_paths
        report.file_path = file_paths.get("pdf") or next(iter(file_paths.values()), None)
        db.commit()
        logger.info(f"[report:{report_id}] Generation complete. Files: {file_paths}")

    except Exception as exc:
        logger.exception(f"[report:{report_id}] Generation failed: {exc}")
        try:
            report = db.query(Report).filter(Report.id == report_id).first()
            if report:
                report.status = "failed"
                report.error_message = str(exc)
                db.commit()
        except Exception:
            pass
        raise
    finally:
        db.close()
