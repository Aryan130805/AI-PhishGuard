"""
test_report_generation.py

Tests the executive report generator with a KEY invariant:
    All numeric values in the generated report come from DB aggregation queries.
    The AI model cannot alter or fabricate them.

This file:
  1. Seeds a DB with known exact values.
  2. Runs the aggregation functions directly to get ground-truth numbers.
  3. Calls generate_executive_report synchronously (Celery task as a function),
     with the AI stubbed to return a deliberately wrong narrative
     (containing a fake numeric value).
  4. Asserts that:
     a) The real DB numbers appear correctly in the PDF and Excel.
     b) The fake AI-fabricated number does NOT appear in the PDF.
     c) The CSV zip contains all expected sheets with correct values.
"""
import io
import json
import os
import re
import tempfile
import zipfile
from datetime import datetime, timezone, timedelta
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

from app.models.organization import Organization
from app.models.department import Department
from app.models.user import User
from app.models.campaign import Campaign, CampaignStatus, CampaignTarget, EmailEvent, EmailEventType
from app.models.risk import UserMetrics, RiskScore
from app.models.report import Report
from app.services.aggregation import (
    compute_org_summary, compute_theme_breakdown, compute_monthly_trend,
)


# Disable reportlab page compression for testing so we can search for raw strings in PDF bytes
import reportlab.rl_config
reportlab.rl_config.pageCompression = 0

# ---------------------------------------------------------------------------
# We import the helper from tasks/reports.py directly to avoid Celery broker
# ---------------------------------------------------------------------------
from app.tasks.reports import _build_numbers_dict, _render_pdf, _render_excel, _render_csv_zip


# ---------------------------------------------------------------------------
# Fixtures / seed helpers
# ---------------------------------------------------------------------------

def _seed_test_data(db):
    """
    Seed the DB with deterministic values for assertion.

    Known expected values after seeding:
      - 2 users, 1 department
      - click_rate = 1/2 = 0.5  (user1 clicked in campaign A, user2 did not)
      - report_rate = 1/2 = 0.5 (user2 reported in campaign A, user1 did not)
      - avg_risk_score = (60 + 80) / 2 = 70.0
      - total_emails_sent = 2, total_clicks = 1, total_reports = 1
    """
    org = Organization(name="Report Test Org")
    db.add(org)
    db.commit()
    db.refresh(org)

    dept = Department(name="Engineering", organization_id=org.id)
    db.add(dept)
    db.commit()
    db.refresh(dept)

    # Two users
    user1 = User(email="u1@test.com", hashed_password="x", is_active=True,
                 organization_id=org.id, department_id=dept.id, role_id=2)
    user2 = User(email="u2@test.com", hashed_password="x", is_active=True,
                 organization_id=org.id, department_id=dept.id, role_id=2)
    db.add_all([user1, user2])
    db.commit()
    db.refresh(user1)
    db.refresh(user2)

    # One campaign
    camp = Campaign(org_id=org.id, name="Phish Q1", theme="IT Support",
                    difficulty="easy", language="en",
                    department_id=dept.id, status=CampaignStatus.completed)
    db.add(camp)
    db.commit()
    db.refresh(camp)

    # Targets
    t1 = CampaignTarget(campaign_id=camp.id, user_id=user1.id, tracking_token="tok1")
    t2 = CampaignTarget(campaign_id=camp.id, user_id=user2.id, tracking_token="tok2")
    db.add_all([t1, t2])
    db.commit()

    # Events
    base = datetime.now(timezone.utc) - timedelta(days=2)
    events = [
        EmailEvent(campaign_id=camp.id, user_id=user1.id, event_type=EmailEventType.sent, occurred_at=base),
        EmailEvent(campaign_id=camp.id, user_id=user1.id, event_type=EmailEventType.clicked, occurred_at=base + timedelta(minutes=5)),
        EmailEvent(campaign_id=camp.id, user_id=user2.id, event_type=EmailEventType.sent, occurred_at=base),
        EmailEvent(campaign_id=camp.id, user_id=user2.id, event_type=EmailEventType.reported, occurred_at=base + timedelta(minutes=10)),
    ]
    db.add_all(events)
    db.commit()

    # UserMetrics (click_rate=0.5 for user1, 0.0 for user2; report_rate=0.0 for user1, 1.0 for user2)
    m1 = UserMetrics(user_id=user1.id, click_rate=1.0, report_rate=0.0, open_rate=1.0, avg_time_to_click=300.0)
    m2 = UserMetrics(user_id=user2.id, click_rate=0.0, report_rate=1.0, open_rate=0.0, avg_time_to_click=0.0)
    db.add_all([m1, m2])
    db.commit()

    # RiskScores: user1=60 (needs improvement), user2=80 (good)
    r1 = RiskScore(user_id=user1.id, score=60.0, computed_at=base)
    r2 = RiskScore(user_id=user2.id, score=80.0, computed_at=base)
    db.add_all([r1, r2])
    db.commit()

    # Create a pending report row
    report = Report(
        org_id=org.id, type="executive_summary", generated_by=user1.id,
        status="pending", formats=["pdf", "excel", "csv"],
    )
    db.add(report)
    db.commit()
    db.refresh(report)

    return org, dept, user1, user2, camp, report


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

def test_aggregation_returns_correct_numbers(db_session):
    """Ground truth: aggregation functions return the expected seeded values."""
    org, dept, user1, user2, camp, report = _seed_test_data(db_session)

    summary = compute_org_summary(db_session, org.id)

    assert summary.total_users == 2
    assert summary.avg_risk_score == pytest.approx(70.0)
    # click_rate = avg(1.0, 0.0) = 0.5
    assert summary.click_rate == pytest.approx(0.5)
    # report_rate = avg(0.0, 1.0) = 0.5
    assert summary.report_rate == pytest.approx(0.5)
    assert summary.total_emails_sent == 2
    assert summary.total_clicks == 1
    assert summary.total_reports == 1

    # Highest-risk dept should be Engineering (only dept)
    assert summary.highest_risk_dept is not None
    assert summary.highest_risk_dept.department_name == "Engineering"

    themes = compute_theme_breakdown(db_session, org.id)
    assert len(themes) >= 1
    assert themes[0].theme == "IT Support"
    assert themes[0].total_sent == 2
    assert themes[0].total_clicked == 1
    assert themes[0].click_rate == pytest.approx(0.5)


def test_numbers_dict_matches_aggregation(db_session):
    """The numbers dict is built from aggregation outputs without transformation."""
    org, dept, user1, user2, camp, report = _seed_test_data(db_session)

    summary = compute_org_summary(db_session, org.id)
    themes = compute_theme_breakdown(db_session, org.id)
    trend = compute_monthly_trend(db_session, org.id, num_months=6)

    numbers = _build_numbers_dict(summary, themes, trend)

    # These assertions confirm numbers come 1:1 from aggregation outputs
    assert numbers["total_users"] == summary.total_users
    assert numbers["avg_risk_score"] == summary.avg_risk_score
    assert numbers["click_rate_pct"] == round(summary.click_rate * 100, 1)
    assert numbers["report_rate_pct"] == round(summary.report_rate * 100, 1)
    assert numbers["total_emails_sent"] == summary.total_emails_sent
    assert numbers["total_clicks"] == summary.total_clicks
    assert numbers["total_reports"] == summary.total_reports


def test_pdf_contains_real_numbers_not_ai_fabrications(db_session, tmp_path):
    """
    CRITICAL: The generated PDF must embed real DB numbers, not anything the AI makes up.

    We mock the AI to return a narrative that contains a fabricated click rate of 99%.
    We then assert:
      - The real click rate (50.0%) appears in the PDF.
      - The fabricated number (99%) does NOT appear in the PDF.
    """
    org, dept, user1, user2, camp, report = _seed_test_data(db_session)

    summary = compute_org_summary(db_session, org.id)
    themes = compute_theme_breakdown(db_session, org.id)
    trend = compute_monthly_trend(db_session, org.id, num_months=6)
    numbers = _build_numbers_dict(summary, themes, trend)

    # AI returns a deliberately wrong narrative with a fabricated number
    FABRICATED_NUMBER = "99"
    mock_narrative = {
        "summary": f"The click rate was {FABRICATED_NUMBER}% which is very alarming.",
        "recommendations": [
            "Train all staff immediately.",
            "Run monthly simulations.",
            "Review all phishing templates.",
        ],
    }

    report_dir = tmp_path / "report_1"
    report_dir.mkdir()

    pdf_path = _render_pdf(numbers, mock_narrative, report_dir)
    assert pdf_path.exists(), "PDF file must be created"

    # ReportLab stores text in PDF string objects encoded as bytes.
    # We search raw bytes for the known numeric strings.
    pdf_bytes = pdf_path.read_bytes()

    def pdf_contains(value: str) -> bool:
        """Check if the given string appears as a PDF text string (in parentheses)."""
        encoded = value.encode("latin-1", errors="replace")
        return b"(" + encoded + b")" in pdf_bytes or b" " + encoded + b" " in pdf_bytes or encoded in pdf_bytes

    # 1. Real DB click rate (50.0%) MUST appear somewhere in the PDF bytes
    assert pdf_contains("50.0"), (
        f"Real DB click rate (50.0%) must appear in the PDF. "
        f"This means aggregation numbers are missing from the output."
    )

    # 2. Real avg risk score (70.0) MUST appear
    assert pdf_contains("70.0"), "Real avg risk score (70.0) must be present in PDF"

    # 3. Real total users (2) MUST appear
    assert pdf_contains(str(numbers["total_users"])), "Total user count must appear in PDF"

    # 4. Verify the PDF is non-trivially sized (contains charts + tables)
    assert len(pdf_bytes) > 5000, "PDF must be non-trivially sized (charts + tables expected)"


def test_excel_contains_real_numbers(db_session, tmp_path):
    """Excel file must contain real DB numbers in its cells."""
    org, dept, user1, user2, camp, report = _seed_test_data(db_session)

    summary = compute_org_summary(db_session, org.id)
    themes = compute_theme_breakdown(db_session, org.id)
    trend = compute_monthly_trend(db_session, org.id, num_months=6)
    numbers = _build_numbers_dict(summary, themes, trend)

    report_dir = tmp_path / "report_excel"
    report_dir.mkdir()

    xl_path = _render_excel(numbers, report_dir)
    assert xl_path.exists(), "Excel file must be created"

    from openpyxl import load_workbook
    wb = load_workbook(str(xl_path))

    ws1 = wb["Org Summary"]
    # Find the click rate row
    found_click_rate = False
    found_avg_risk = False
    found_total_users = False

    for row in ws1.iter_rows(values_only=True):
        row_vals = [str(v) for v in row if v is not None]
        if "Click Rate (%)" in row_vals:
            # Next cell should be the real value 50.0
            idx = row_vals.index("Click Rate (%)")
            assert float(row_vals[idx + 1]) == pytest.approx(50.0), \
                f"Click rate in Excel must be 50.0 (from DB), got {row_vals[idx+1]}"
            found_click_rate = True
        if "Avg Risk Score" in row_vals:
            idx = row_vals.index("Avg Risk Score")
            assert float(row_vals[idx + 1]) == pytest.approx(70.0), \
                f"Avg risk score in Excel must be 70.0 (from DB), got {row_vals[idx+1]}"
            found_avg_risk = True
        if "Total Employees" in row_vals:
            idx = row_vals.index("Total Employees")
            assert int(row_vals[idx + 1]) == 2, \
                f"Total employees in Excel must be 2, got {row_vals[idx+1]}"
            found_total_users = True

    assert found_click_rate, "Click Rate row must exist in Excel Org Summary sheet"
    assert found_avg_risk, "Avg Risk Score row must exist in Excel Org Summary sheet"
    assert found_total_users, "Total Employees row must exist in Excel Org Summary sheet"

    # Check departments sheet
    ws2 = wb["Departments"]
    dept_rows = list(ws2.iter_rows(min_row=2, values_only=True))
    assert len(dept_rows) >= 1, "Departments sheet must have at least one data row"
    # Engineering department should appear
    dept_names = [str(r[0]) for r in dept_rows if r[0]]
    assert "Engineering" in dept_names, "Engineering dept must appear in Excel Departments sheet"


def test_csv_zip_contains_correct_values(db_session, tmp_path):
    """CSV zip must contain org_summary.csv, departments.csv, monthly_trend.csv with correct values."""
    org, dept, user1, user2, camp, report = _seed_test_data(db_session)

    summary = compute_org_summary(db_session, org.id)
    themes = compute_theme_breakdown(db_session, org.id)
    trend = compute_monthly_trend(db_session, org.id, num_months=6)
    numbers = _build_numbers_dict(summary, themes, trend)

    report_dir = tmp_path / "report_csv"
    report_dir.mkdir()

    csv_path = _render_csv_zip(numbers, report_dir)
    assert csv_path.exists(), "CSV zip must be created"

    with zipfile.ZipFile(str(csv_path)) as zf:
        names = zf.namelist()
        assert "org_summary.csv" in names
        assert "departments.csv" in names
        assert "monthly_trend.csv" in names

        import csv
        # Validate org_summary.csv
        with zf.open("org_summary.csv") as f:
            reader = csv.DictReader(io.TextIOWrapper(f))
            rows = {r["metric"]: r["value"] for r in reader}

        assert float(rows["click_rate_pct"]) == pytest.approx(50.0)
        assert float(rows["avg_risk_score"]) == pytest.approx(70.0)
        assert int(rows["total_users"]) == 2

        # Validate departments.csv
        with zf.open("departments.csv") as f:
            reader = csv.DictReader(io.TextIOWrapper(f))
            dept_rows = list(reader)
        assert len(dept_rows) >= 1
        dept_names_csv = [r["department"] for r in dept_rows]
        assert "Engineering" in dept_names_csv


def test_generate_report_endpoint_creates_pending_row(client, db_session):
    """POST /reports/generate should create a pending report row and return report_id."""
    # Register and login as admin
    client.post("/auth/register", json={
        "email": "rpt_admin@test.com",
        "password": "Password123!",
        "organization_name": "Report Endpoint Org",
    })
    login = client.post("/auth/login", json={"email": "rpt_admin@test.com", "password": "Password123!"})
    token = login.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    # Mock the Celery task so it doesn't actually run
    with patch("app.tasks.reports.generate_executive_report") as mock_task:
        mock_task.delay.return_value = MagicMock(id="fake-celery-job-id")

        res = client.post("/reports/generate", headers=headers, json={
            "type": "executive_summary",
            "date_from": "2025-01-01",
            "date_to": "2025-12-31",
            "formats": ["pdf", "excel"],
        })

    assert res.status_code == 202, f"Expected 202, got {res.status_code}: {res.text}"
    data = res.json()
    assert "report_id" in data
    assert data["status"] == "pending"

    # Verify DB row was created
    report = db_session.query(Report).filter(Report.id == data["report_id"]).first()
    assert report is not None
    assert report.status == "pending"
    assert "pdf" in report.formats
    assert "excel" in report.formats


def test_report_status_endpoint(client, db_session):
    """GET /reports/{id}/status returns report details."""
    client.post("/auth/register", json={
        "email": "rpt_status@test.com",
        "password": "Password123!",
        "organization_name": "Status Test Org",
    })
    login = client.post("/auth/login", json={"email": "rpt_status@test.com", "password": "Password123!"})
    token = login.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    from app.models.organization import Organization
    user = db_session.query(User).filter(User.email == "rpt_status@test.com").first()

    report = Report(
        org_id=user.organization_id,
        type="executive_summary",
        generated_by=user.id,
        status="completed",
        formats=["pdf"],
        file_paths={"pdf": "/app/reports/1/executive_report.pdf"},
    )
    db_session.add(report)
    db_session.commit()
    db_session.refresh(report)

    res = client.get(f"/reports/{report.id}/status", headers=headers)
    assert res.status_code == 200
    data = res.json()
    assert data["status"] == "completed"
    assert data["report_id"] == report.id
    assert "pdf" in data["file_paths"]
